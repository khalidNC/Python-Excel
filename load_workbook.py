import openpyxl

# Define a class for loading a workbook
class LoadWorkbook:
  # Define constractor and the class with one attribute xlsx_file 
  def __init__(self, xlsx_file):
    self.xlsx_file = xlsx_file

  ''' 
  Define a function loadWorkbook that returns wb object which takes a xcel file and loads using 
  built-in method load_workbook() from openpyxl module
  '''
  def loadWorkbook(self):
    wb = openpyxl.load_workbook(self.xlsx_file)
    return wb
  

# Create class object and it returns loader
loder = LoadWorkbook("transactions.xlsx")

# Call the loadWorkbook() method and it returns wb object
wb = loder.loadWorkbook()
