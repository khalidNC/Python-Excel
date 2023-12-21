import openpyxl

# Load a exiting workbook using load_workbook() method and it returns wb object
wb = openpyxl.load_workbook("transactions.xlsx")

# wb object has an attribute sheetnames to get sheet'a name and print in terminal
print(wb.sheetnames)

# Access the sheet using square brackets like dictionary and it returns sheet object
sheet = wb["Sheet1"]

# Creating a new sheet using create_sheet() method and pass the new sheet name in string
# and set position index for the first position by passing 0 as second argument
wb.create_sheet("home", 0)

# Save the workbook in new xlsx file
wb.save("tr.xlsx")

# Load the newly created wb and it returns wb1 object
wb1 = openpyxl.load_workbook("tr.xlsx")

# print the sheet names
print(wb1.sheetnames)

# remove home sheet then save the changes to a newly created sheet
sheet_remove = wb1["home"]
wb1.remove(sheet_remove)
wb1.save("ttt.xlsx")







